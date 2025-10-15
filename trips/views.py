from django.shortcuts import render
from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db import transaction
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import models
import requests
import json
import math
from .models import Vehicle, Trip, RestStop, TripWaypoint, VehicleAssignment
from .serializers import (
    VehicleSerializer, TripSerializer, TripCreateSerializer, TripUpdateSerializer,
    RestStopSerializer, TripWaypointSerializer, TripPlanningSerializer,
    RouteCalculationSerializer, VehicleAssignmentSerializer
)
from .services import HOSCalculator, NominatimService
from accounts.views import IsFleetManagerOrAdmin
from accounts.models import User
from accounts.serializers import UserSerializer
import logging
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO

logger = logging.getLogger(__name__)

class VehicleListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer les véhicules"""
    
    queryset = Vehicle.objects.select_related('company', 'current_driver')
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.is_admin():
            return Vehicle.objects.select_related('company', 'current_driver').all()
        else:
            return Vehicle.objects.filter(company=user.company).select_related('company', 'current_driver')
    
    def perform_create(self, serializer):
        if self.request.user.is_admin():
            serializer.save()
        else:
            serializer.save(company=self.request.user.company)

class VehicleDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour consulter, modifier et supprimer un véhicule"""
    
    queryset = Vehicle.objects.select_related('company', 'current_driver')
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_object(self):
        obj = super().get_object()
        user = self.request.user
        
        if user.is_admin() or obj.company == user.company:
            return obj
        else:
            raise permissions.PermissionDenied("Not authorized to access this vehicle")

class VehicleAssignDriverView(APIView):
    """Vue pour assigner un conducteur à un véhicule"""
    
    permission_classes = [permissions.IsAuthenticated, IsFleetManagerOrAdmin]
    
    def post(self, request, vehicle_id):
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)
            
            # Vérifier les permissions
            user = request.user
            if not user.is_admin() and vehicle.company != user.company:
                raise permissions.PermissionDenied("Not authorized to modify this vehicle")
            
            driver_id = request.data.get('driver_id')
            if not driver_id:
                return Response({
                    'error': 'Driver ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                driver = User.objects.get(id=driver_id, user_type='DRIVER')
                
                # Vérifier que le conducteur appartient à la même compagnie
                if not user.is_admin() and driver.company != user.company:
                    raise permissions.PermissionDenied("Driver must belong to the same company")
                
            except User.DoesNotExist:
                return Response({
                    'error': 'Driver not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Créer une nouvelle attribution
            with transaction.atomic():
                # Terminer l'attribution actuelle si elle existe
                if vehicle.current_driver:
                    current_assignment = VehicleAssignment.objects.filter(
                        vehicle=vehicle,
                        driver=vehicle.current_driver,
                        end_date__isnull=True
                    ).first()
                    
                    if current_assignment:
                        current_assignment.end_date = timezone.now()
                        current_assignment.save()
                
                # Créer la nouvelle attribution
                new_assignment = VehicleAssignment.objects.create(
                    vehicle=vehicle,
                    driver=driver,
                    assigned_by=user,
                    start_date=timezone.now(),
                    notes=request.data.get('notes', '')
                )
                
                # Mettre à jour le véhicule
                vehicle.current_driver = driver
                vehicle.save()
            
            return Response({
                'success': True,
                'message': f'Vehicle {vehicle.license_plate} assigned to {driver.get_full_name()}',
                'assignment': VehicleAssignmentSerializer(new_assignment).data,
                'vehicle': VehicleSerializer(vehicle).data
            })
            
        except Vehicle.DoesNotExist:
            return Response({
                'error': 'Vehicle not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error assigning driver to vehicle: {str(e)}")
            return Response({
                'error': 'Internal error during assignment'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class VehicleUnassignDriverView(APIView):
    """Vue pour désassigner un conducteur d'un véhicule"""
    
    permission_classes = [permissions.IsAuthenticated, IsFleetManagerOrAdmin]
    
    def post(self, request, vehicle_id):
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)
            
            # Vérifier les permissions
            user = request.user
            if not user.is_admin() and vehicle.company != user.company:
                raise permissions.PermissionDenied("Not authorized to modify this vehicle")
            
            if not vehicle.current_driver:
                return Response({
                    'error': 'No driver currently assigned to this vehicle'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                # Terminer l'attribution actuelle
                current_assignment = VehicleAssignment.objects.filter(
                    vehicle=vehicle,
                    driver=vehicle.current_driver,
                    end_date__isnull=True
                ).first()
                
                if current_assignment:
                    current_assignment.end_date = timezone.now()
                    current_assignment.ended_by = user
                    current_assignment.end_reason = request.data.get('notes', current_assignment.notes)
                    current_assignment.save()
                
                # Retirer le conducteur du véhicule
                previous_driver = vehicle.current_driver
                vehicle.current_driver = None
                vehicle.save()
            
            return Response({
                'success': True,
                'message': f'Driver {previous_driver.get_full_name()} unassigned from vehicle {vehicle.license_plate}',
                'vehicle': VehicleSerializer(vehicle).data
            })
            
        except Vehicle.DoesNotExist:
            return Response({
                'error': 'Vehicle not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error unassigning driver from vehicle: {str(e)}")
            return Response({
                'error': 'Internal error during unassignment'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class VehicleAssignmentHistoryView(generics.ListAPIView):
    """Vue pour consulter l'historique d'attribution d'un véhicule"""
    
    serializer_class = VehicleAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        vehicle_id = self.kwargs.get('vehicle_id')
        user = self.request.user
        
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)
            
            # Vérifier les permissions
            if not user.is_admin() and vehicle.company != user.company:
                raise permissions.PermissionDenied("Not authorized to access this vehicle's history")
            
            return VehicleAssignment.objects.filter(vehicle=vehicle).select_related(
                'driver', 'assigned_by', 'unassigned_by'
            ).order_by('-start_time')
            
        except Vehicle.DoesNotExist:
            return VehicleAssignment.objects.none()

class AvailableDriversForVehicleView(generics.ListAPIView):
    """Vue pour lister les conducteurs disponibles pour un véhicule"""
    
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated, IsFleetManagerOrAdmin]
    
    def get_queryset(self):
        vehicle_id = self.kwargs.get('vehicle_id')
        user = self.request.user
        
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)
            
            # Vérifier les permissions
            if not user.is_admin() and vehicle.company != user.company:
                raise permissions.PermissionDenied("Not authorized to access this vehicle")
            
            # Retourner les conducteurs de la même compagnie qui ne sont pas en voyage actif
            if user.is_admin():
                available_drivers = User.objects.filter(
                    user_type='DRIVER',
                    is_active=True
                ).exclude(
                    trips__status='IN_PROGRESS'
                ).distinct()
            else:
                available_drivers = User.objects.filter(
                    user_type='DRIVER',
                    company=vehicle.company,
                    is_active=True
                ).exclude(
                    trips__status='IN_PROGRESS'
                ).distinct()
            
            return available_drivers
            
        except Vehicle.DoesNotExist:
            return User.objects.none()

class TripListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer les voyages"""
    
    queryset = Trip.objects.select_related('driver', 'vehicle', 'created_by').prefetch_related('rest_stops', 'waypoints')
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TripCreateSerializer
        return TripSerializer
    
    def get_queryset(self):
        user = self.request.user
        queryset = Trip.objects.select_related('driver', 'vehicle', 'created_by').prefetch_related('rest_stops', 'waypoints')
        
        # Filtrage selon les permissions
        if user.user_type == 'ADMIN':
            filtered_queryset = queryset
        elif user.user_type == 'FLEET_MANAGER':
            filtered_queryset = queryset.filter(driver__company=user.company)
        else:  # DRIVER
            filtered_queryset = queryset.filter(driver=user)
        
        # Filtres supplémentaires via paramètres de requête
        status = self.request.query_params.get('status')
        if status:
            filtered_queryset = filtered_queryset.filter(status=status)
        
        date_from = self.request.query_params.get('date_from')
        if date_from:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                filtered_queryset = filtered_queryset.filter(planned_departure__date__gte=date_obj)
            except ValueError:
                pass
        
        date_to = self.request.query_params.get('date_to')
        if date_to:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                filtered_queryset = filtered_queryset.filter(planned_departure__date__lte=date_obj)
            except ValueError:
                pass
        
        priority = self.request.query_params.get('priority')
        if priority:
            filtered_queryset = filtered_queryset.filter(priority=priority)
        
        vehicle_id = self.request.query_params.get('vehicle')
        if vehicle_id:
            filtered_queryset = filtered_queryset.filter(vehicle_id=vehicle_id)
        
        driver_id = self.request.query_params.get('driver')
        if driver_id:
            filtered_queryset = filtered_queryset.filter(driver_id=driver_id)
        
        return filtered_queryset.order_by('-planned_departure')
    
    def perform_create(self, serializer):
        user = self.request.user
        
        # Vérifier les permissions
        if user.user_type not in ['FLEET_MANAGER', 'DRIVER']:
            raise permissions.PermissionDenied("Seuls les gestionnaires de flotte et les conducteurs peuvent créer des voyages.")
        
        # Ajouter le créateur
        serializer.save(created_by=user)

class TripDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour consulter, modifier et supprimer un voyage"""
    
    queryset = Trip.objects.select_related('driver', 'vehicle', 'created_by', 'updated_by').prefetch_related('rest_stops', 'waypoints')
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return TripUpdateSerializer
        return TripSerializer
    
    def get_object(self):
        obj = super().get_object()
        user = self.request.user
        
        # Vérifier les permissions d'accès
        if (user.user_type == 'ADMIN' or 
            (user.user_type == 'FLEET_MANAGER' and obj.driver.company == user.company) or
            obj.driver == user):
            return obj
        else:
            raise permissions.PermissionDenied("Non autorisé à accéder à ce voyage")
    
    def perform_update(self, serializer):
        # Ajouter l'utilisateur qui modifie
        serializer.save(updated_by=self.request.user)

class TripStartView(APIView):
    """Vue pour démarrer un voyage"""
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, trip_id):
        try:
            user = request.user
            
            try:
                trip = Trip.objects.get(id=trip_id)
            except Trip.DoesNotExist:
                return Response({
                    'error': 'Voyage non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Vérifier les permissions
            if not (user.user_type == 'ADMIN' or trip.driver == user or 
                    (user.user_type == 'FLEET_MANAGER' and trip.driver.company == user.company)):
                return Response({
                    'error': 'Permission refusée'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Vérifier si le voyage peut être démarré
            can_start, reason = trip.can_be_started()
            if not can_start:
                return Response({
                    'error': f'Impossible de démarrer le voyage : {reason}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Récupérer la position actuelle optionnelle
            current_latitude = request.data.get('current_latitude')
            current_longitude = request.data.get('current_longitude')
            
            # Démarrer le voyage
            try:
                trip.start_trip(current_latitude, current_longitude)
                
                return Response({
                    'success': True,
                    'message': 'Voyage démarré avec succès',
                    'trip': TripSerializer(trip).data
                })
                
            except ValueError as e:
                return Response({
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.error(f"Erreur lors du démarrage du voyage: {str(e)}")
            return Response({
                'error': 'Erreur interne lors du démarrage du voyage'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TripCompleteView(APIView):
    """Vue pour terminer un voyage"""
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, trip_id):
        try:
            user = request.user
            
            try:
                trip = Trip.objects.get(id=trip_id)
            except Trip.DoesNotExist:
                return Response({
                    'error': 'Voyage non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Vérifier les permissions
            if not (user.user_type == 'ADMIN' or trip.driver == user or 
                    (user.user_type == 'FLEET_MANAGER' and trip.driver.company == user.company)):
                return Response({
                    'error': 'Permission refusée'
                }, status=status.HTTP_403_FORBIDDEN)
            
            if trip.status != 'IN_PROGRESS':
                return Response({
                    'error': f'Impossible de terminer le voyage avec le statut: {trip.get_status_display()}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Récupérer la position finale optionnelle
            final_latitude = request.data.get('final_latitude')
            final_longitude = request.data.get('final_longitude')
            
            # Ajouter des notes du conducteur si fournies
            driver_notes = request.data.get('driver_notes', '')
            if driver_notes:
                trip.driver_notes = f"{trip.driver_notes}\n\n{driver_notes}".strip()
            
            # Terminer le voyage
            try:
                trip.complete_trip(final_latitude, final_longitude)
                
                return Response({
                    'success': True,
                    'message': 'Voyage terminé avec succès',
                    'trip': TripSerializer(trip).data
                })
                
            except ValueError as e:
                return Response({
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.error(f"Erreur lors de la finalisation du voyage: {str(e)}")
            return Response({
                'error': 'Erreur interne lors de la finalisation du voyage'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TripCancelView(APIView):
    """Vue pour annuler un voyage"""
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, trip_id):
        try:
            user = request.user
            
            try:
                trip = Trip.objects.get(id=trip_id)
            except Trip.DoesNotExist:
                return Response({
                    'error': 'Voyage non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Vérifier les permissions
            if not (user.user_type == 'ADMIN' or trip.driver == user or 
                    (user.user_type == 'FLEET_MANAGER' and trip.driver.company == user.company)):
                return Response({
                    'error': 'Permission refusée'
                }, status=status.HTTP_403_FORBIDDEN)
            
            if trip.status == 'COMPLETED':
                return Response({
                    'error': 'Un voyage terminé ne peut pas être annulé'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Récupérer la raison d'annulation
            reason = request.data.get('reason', 'Annulation demandée par l\'utilisateur')
            
            # Annuler le voyage
            try:
                trip.cancel_trip(reason)
                
                return Response({
                    'success': True,
                    'message': 'Voyage annulé avec succès',
                    'trip': TripSerializer(trip).data
                })
                
            except ValueError as e:
                return Response({
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.error(f"Erreur lors de l'annulation du voyage: {str(e)}")
            return Response({
                'error': 'Erreur interne lors de l\'annulation du voyage'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TripUpdatePositionView(APIView):
    """Vue pour mettre à jour la position GPS d'un voyage en cours"""
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, trip_id):
        try:
            user = request.user
            
            # Seuls les conducteurs peuvent mettre à jour leur position
            if user.user_type != 'DRIVER':
                return Response({
                    'error': 'Seuls les conducteurs peuvent mettre à jour la position'
                }, status=status.HTTP_403_FORBIDDEN)
            
            try:
                trip = Trip.objects.get(id=trip_id, driver=user, status='IN_PROGRESS')
            except Trip.DoesNotExist:
                return Response({
                    'error': 'Voyage actif non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
            
            latitude = request.data.get('latitude')
            longitude = request.data.get('longitude')
            
            if not latitude or not longitude:
                return Response({
                    'error': 'Latitude et longitude requises'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                lat = float(latitude)
                lng = float(longitude)
                
                # Valider les coordonnées
                if not (-90 <= lat <= 90):
                    raise ValueError("Latitude invalide")
                if not (-180 <= lng <= 180):
                    raise ValueError("Longitude invalide")
                
            except (ValueError, TypeError):
                return Response({
                    'error': 'Valeurs de coordonnées invalides'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Mettre à jour la position
            try:
                trip.update_position(lat, lng)
                
                # Calculer la distance restante
                distance_remaining = trip.get_distance_remaining()
                
                # Vérifier si arrivé à destination (moins de 100m)
                is_arrived = distance_remaining and distance_remaining < 0.1
                
                response_data = {
                    'success': True,
                    'position_updated': True,
                    'current_position': {
                        'latitude': lat,
                        'longitude': lng,
                        'timestamp': timezone.now().isoformat()
                    },
                    'navigation': {
                        'remaining_distance_km': round(distance_remaining, 2) if distance_remaining else None,
                        'is_arrived': is_arrived,
                        'progress_percentage': trip.progress_percentage
                    }
                }
                
                # Proposer de terminer le voyage si arrivé
                if is_arrived:
                    response_data['arrival_detected'] = True
                    response_data['message'] = 'Arrivée détectée - Vous pouvez terminer le voyage'
                
                return Response(response_data)
                
            except ValueError as e:
                return Response({
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour de position: {str(e)}")
            return Response({
                'error': 'Erreur lors de la mise à jour de la position'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TripPlanningView(APIView):
    """Vue pour la planification automatique de voyages avec calculs HOS"""
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        serializer = TripPlanningSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                # Utiliser le HOSCalculator pour planifier le voyage
                calculator = HOSCalculator(serializer.validated_data)
                trip_schedule = calculator.calculate_trip_schedule()
                
                return Response({
                    'success': True,
                    'trip_schedule': trip_schedule,
                    'message': 'Trip planning completed successfully'
                })
                
            except ValueError as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                logger.error(f"Trip planning error: {str(e)}")
                return Response({
                    'success': False,
                    'error': 'Internal error during trip planning'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class RouteCalculationView(APIView):
    """Vue pour le calcul d'itinéraires optimisés"""
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        serializer = RouteCalculationSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                waypoints = serializer.validated_data['waypoints']
                
                # Créer des données temporaires pour le calculateur
                temp_trip_data = {
                    'current_location': 'Start',
                    'pickup_location': 'Waypoint 1',
                    'dropoff_location': 'End',
                    'current_cycle_hours': 0,
                    'planned_start_time': timezone.now().isoformat()
                }
                
                calculator = HOSCalculator(temp_trip_data)
                route_data = calculator.get_route_from_api([
                    f"{wp['lat']},{wp['lng']}" for wp in waypoints
                ])
                
                return Response({
                    'success': True,
                    'route_data': route_data
                })
                
            except Exception as e:
                logger.error(f"Route calculation error: {str(e)}")
                return Response({
                    'success': False,
                    'error': 'Error calculating route'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class RestStopListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer les arrêts de repos"""
    
    serializer_class = RestStopSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        trip_id = self.kwargs.get('trip_id')
        return RestStop.objects.filter(trip_id=trip_id).order_by('planned_start_time')
    
    def perform_create(self, serializer):
        trip_id = self.kwargs.get('trip_id')
        try:
            trip = Trip.objects.get(id=trip_id)
            
            # Vérifier les permissions
            user = self.request.user
            if not (user.is_admin() or 
                   (user.is_fleet_manager() and trip.driver.company == user.company) or
                   trip.driver == user):
                raise permissions.PermissionDenied("Not authorized to modify this trip")
            
            serializer.save(trip=trip)
            
        except Trip.DoesNotExist:
            raise generics.Http404("Trip not found")

class RestStopDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour consulter, modifier et supprimer un arrêt de repos"""
    
    queryset = RestStop.objects.all()
    serializer_class = RestStopSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_object(self):
        obj = super().get_object()
        user = self.request.user
        
        if (user.is_admin() or 
            (user.is_fleet_manager() and obj.trip.driver.company == user.company) or
            obj.trip.driver == user):
            return obj
        else:
            raise permissions.PermissionDenied("Not authorized to access this rest stop")

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def search_locations(request):
    """API pour l'autocomplétion des lieux avec Nominatim"""
    try:
        query = request.GET.get('q', '').strip()
        if len(query) < 3:
            return Response({
                'suggestions': [],
                'message': 'Veuillez saisir au moins 3 caractères'
            })
        
        nominatim_service = NominatimService()
        suggestions = nominatim_service.search_address(query, limit=8)
        
        return Response({
            'suggestions': suggestions,
            'count': len(suggestions)
        })
        
    except Exception as e:
        logger.error(f"Error in location search: {str(e)}")
        return Response({
            'error': 'Erreur lors de la recherche d\'adresses',
            'suggestions': []
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def calculate_route_with_navigation(request):
    """Calculer un itinéraire complet avec données de navigation GPS"""
    try:
        origin = request.data.get('origin')
        destination = request.data.get('destination')
        waypoints = request.data.get('waypoints', [])
        
        if not origin or not destination:
            return Response({
                'error': 'Origin et destination sont requis'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        nominatim_service = NominatimService()
        
        # Calculer l'itinéraire avec OpenRouteService ou équivalent
        route_data = nominatim_service.calculate_route(origin, destination, waypoints)
        
        if not route_data:
            return Response({
                'error': 'Impossible de calculer l\'itinéraire'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': True,
            'route': route_data,
            'navigation': {
                'total_distance_km': route_data.get('distance_km', 0),
                'total_duration_minutes': route_data.get('duration_minutes', 0),
                'polyline': route_data.get('polyline', ''),
                'turn_by_turn': route_data.get('instructions', []),
                'bbox': route_data.get('bbox', [])
            }
        })
        
    except Exception as e:
        logger.error(f"Error calculating route: {str(e)}")
        return Response({
            'error': 'Erreur lors du calcul de l\'itinéraire'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def get_current_location_address(request):
    """Obtenir l'adresse à partir des coordonnées GPS (géocodage inverse)"""
    try:
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        
        if not latitude or not longitude:
            return Response({
                'error': 'Latitude et longitude requises'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        nominatim_service = NominatimService()
        address_data = nominatim_service.reverse_geocode(float(latitude), float(longitude))
        
        if not address_data:
            return Response({
                'error': 'Impossible de déterminer l\'adresse'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': True,
            'address': address_data
        })
        
    except Exception as e:
        logger.error(f"Error in reverse geocoding: {str(e)}")
        return Response({
            'error': 'Erreur lors de la géolocalisation inverse'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def start_trip_with_navigation(request, trip_id):
    """Démarrer un voyage avec navigation GPS complète"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent démarrer la navigation'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trip = Trip.objects.get(id=trip_id, driver=user)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé ou non assigné'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if trip.status != 'PLANNED':
            return Response({
                'error': f'Impossible de démarrer le voyage avec le statut: {trip.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Récupérer la position actuelle du conducteur
        current_latitude = request.data.get('current_latitude')
        current_longitude = request.data.get('current_longitude')
        
        if not current_latitude or not current_longitude:
            return Response({
                'error': 'Position actuelle requise pour la navigation'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            lat = float(current_latitude)
            lng = float(current_longitude)
        except (ValueError, TypeError):
            return Response({
                'error': 'Coordonnées GPS invalides'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Calculer l'itinéraire complet
        nominatim_service = NominatimService()
        
        # Géocoder la destination si pas encore fait
        if not hasattr(trip, 'destination_lat') or not trip.destination_lat:
            dest_coords = nominatim_service.geocode_address(trip.destination)
            if dest_coords:
                trip.destination_lat = dest_coords['lat']
                trip.destination_lng = dest_coords['lng']
        
        # Calculer la route
        origin = {'lat': lat, 'lng': lng}
        destination = {
            'lat': trip.destination_lat or 0,
            'lng': trip.destination_lng or 0,
            'address': trip.destination
        }
        
        route_data = nominatim_service.calculate_route(origin, destination)
        if not route_data:
            return Response({
                'error': 'Impossible de calculer l\'itinéraire de navigation'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Démarrer le voyage avec navigation
        trip.status = 'IN_PROGRESS'
        trip.start_time = timezone.now()
        trip.current_lat = lat
        trip.current_lng = lng
        trip.route_data = json.dumps(route_data)
        trip.estimated_distance_km = route_data.get('distance_km', 0)
        trip.estimated_duration_minutes = route_data.get('duration_minutes', 0)
        trip.save()
        
        # Mettre à jour le véhicule
        if hasattr(user, 'current_vehicle') and user.current_vehicle:
            vehicle = user.current_vehicle
            vehicle.current_latitude = lat
            vehicle.current_longitude = lng
            vehicle.last_location_update = timezone.now()
            vehicle.save()
        
        return Response({
            'success': True,
            'message': 'Navigation démarrée avec succès',
            'trip': TripSerializer(trip).data,
            'navigation': {
                'status': 'ACTIVE',
                'current_position': {'lat': lat, 'lng': lng},
                'destination': destination,
                'route': route_data
            },
            'redirect_url': f'/tracking/{trip_id}'
        })
        
    except Exception as e:
        logger.error(f"Erreur lors du démarrage de la navigation: {str(e)}")
        return Response({
            'error': 'Erreur lors du démarrage de la navigation'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def update_navigation_position(request, trip_id):
    """Mettre à jour la position GPS pendant la navigation"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent mettre à jour la position'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trip = Trip.objects.get(id=trip_id, driver=user, status='IN_PROGRESS')
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage actif non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        bearing = request.data.get('bearing', 0)
        speed = request.data.get('speed', 0)
        
        if not latitude or not longitude:
            return Response({
                'error': 'Latitude et longitude requises'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            lat = float(latitude)
            lng = float(longitude)
            bearing = float(bearing) if bearing else 0
            speed = float(speed) if speed else 0
        except (ValueError, TypeError):
            return Response({
                'error': 'Valeurs de coordonnées invalides'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Calculer la distance restante vers la destination
        nominatim_service = NominatimService()
        remaining_distance = 0
        
        if hasattr(trip, 'destination_lat') and trip.destination_lat:
            remaining_distance = nominatim_service.calculate_distance(
                lat, lng, trip.destination_lat, trip.destination_lng
            )
        
        # Mettre à jour la position
        trip.current_lat = lat
        trip.current_lng = lng
        trip.save()
        
        # Mettre à jour le véhicule
        if hasattr(user, 'current_vehicle') and user.current_vehicle:
            vehicle = user.current_vehicle
            vehicle.current_latitude = lat
            vehicle.current_longitude = lng
            vehicle.last_location_update = timezone.now()
            vehicle.save()
        
        # Vérifier si arrivé à destination (moins de 100m)
        is_arrived = remaining_distance < 0.1  # 100 mètres
        
        response_data = {
            'success': True,
            'position_updated': True,
            'current_position': {
                'lat': lat,
                'lng': lng,
                'bearing': bearing,
                'speed_kmh': speed,
                'timestamp': timezone.now().isoformat()
            },
            'navigation': {
                'remaining_distance_km': round(remaining_distance, 2),
                'is_arrived': is_arrived,
                'estimated_arrival': trip.estimated_duration_minutes
            }
        }
        
        # Auto-complétion du voyage si arrivé
        if is_arrived and trip.status == 'IN_PROGRESS':
            trip.status = 'COMPLETED'
            trip.end_time = timezone.now()
            trip.actual_distance_km = trip.estimated_distance_km
            trip.save()
            
            response_data['trip_completed'] = True
            response_data['message'] = 'Voyage terminé automatiquement - Arrivé à destination'
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour de position: {str(e)}")
        return Response({
            'error': 'Erreur lors de la mise à jour de la position'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def dashboard_stats(request):
    """Statistiques du tableau de bord"""
    try:
        user = request.user
        
        # Filtrer selon les permissions
        if user.is_admin():
            vehicles = Vehicle.objects.all()
            trips = Trip.objects.all()
        elif user.is_fleet_manager():
            vehicles = Vehicle.objects.filter(company=user.company)
            trips = Trip.objects.filter(driver__company=user.company)
        else:
            vehicles = Vehicle.objects.filter(current_driver=user)
            trips = Trip.objects.filter(driver=user)
        
        # Calculer les statistiques
        stats = {
            'total_vehicles': vehicles.count(),
            'active_vehicles': vehicles.filter(operational_status='IN_USE').count(),
            'available_vehicles': vehicles.filter(operational_status='AVAILABLE').count(),
            'maintenance_vehicles': vehicles.filter(operational_status='MAINTENANCE').count(),
            
            'total_trips': trips.count(),
            'active_trips': trips.filter(status='IN_PROGRESS').count(),
            'completed_trips': trips.filter(status='COMPLETED').count(),
            'planned_trips': trips.filter(status='PLANNED').count(),
            
            'recent_trips': TripSerializer(trips.order_by('-created_at')[:5], many=True).data
        }
        
        return Response(stats)
        
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des statistiques'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def hos_status(request):
    """Statut HOS du conducteur"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs ont un statut HOS'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Calculer le statut HOS actuel
        hos_data = {
            'driver_id': user.id,
            'current_status': 'off_duty',
            'drive_time_used': 0,
            'duty_time_used': 0,
            'cycle_time_used': 0,
            'remaining_drive_time': 11 * 60,  # 11 heures en minutes
            'remaining_duty_time': 14 * 60,   # 14 heures en minutes
            'last_break': None,
            'violations': []
        }
        
        return Response(hos_data)
        
    except Exception as e:
        logger.error(f"Error getting HOS status: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération du statut HOS'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_dashboard_data(request):
    """Données complètes du tableau de bord"""
    try:
        stats_response = dashboard_stats(request)
        hos_response = hos_status(request) if request.user.is_driver() else None
        
        dashboard_data = {
            'stats': stats_response.data,
            'hos_status': hos_response.data if hos_response and hos_response.status_code == 200 else None,
            'timestamp': timezone.now().isoformat()
        }
        
        return Response(dashboard_data)
        
    except Exception as e:
        logger.error(f"Error getting dashboard data: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des données du tableau de bord'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def assign_driver_to_vehicle(request):
    """Assigner un conducteur à un véhicule"""
    try:
        vehicle_id = request.data.get('vehicle_id')
        driver_id = request.data.get('driver_id')
        
        if not vehicle_id or not driver_id:
            return Response({
                'error': 'vehicle_id et driver_id sont requis'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)
            driver = User.objects.get(id=driver_id)
        except (Vehicle.DoesNotExist, User.DoesNotExist):
            return Response({
                'error': 'Véhicule ou conducteur non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier les permissions
        user = request.user
        if not (user.is_admin() or (user.is_fleet_manager() and vehicle.company == user.company)):
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Désassigner le véhicule précédent du conducteur
        if hasattr(driver, 'current_vehicle') and driver.current_vehicle:
            old_vehicle = driver.current_vehicle
            old_vehicle.current_driver = None
            old_vehicle.save()
        
        # Désassigner l'ancien conducteur du véhicule
        if vehicle.current_driver:
            old_driver = vehicle.current_driver
            # Note: Pas besoin de modifier l'utilisateur car la relation se fait via le véhicule
        
        # Nouvelle assignation
        vehicle.current_driver = driver
        vehicle.save()
        
        return Response({
            'success': True,
            'message': f'Conducteur {driver.get_full_name()} assigné au véhicule {vehicle.vehicle_number}',
            'vehicle': VehicleSerializer(vehicle).data
        })
        
    except Exception as e:
        logger.error(f"Error assigning driver to vehicle: {str(e)}")
        return Response({
            'error': 'Erreur lors de l\'assignation'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def unassign_driver_from_vehicle(request):
    """Désassigner un conducteur d'un véhicule"""
    try:
        vehicle_id = request.data.get('vehicle_id')
        
        if not vehicle_id:
            return Response({
                'error': 'vehicle_id requis'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)
        except Vehicle.DoesNotExist:
            return Response({
                'error': 'Véhicule non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier les permissions
        user = request.user
        if not (user.is_admin() or (user.is_fleet_manager() and vehicle.company == user.company)):
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if not vehicle.current_driver:
            return Response({
                'error': 'Aucun conducteur assigné à ce véhicule'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        driver = vehicle.current_driver
        
        # Désassignation
        vehicle.current_driver = None
        vehicle.save()
        
        return Response({
            'success': True,
            'message': f'Conducteur {driver.get_full_name()} désassigné du véhicule {vehicle.vehicle_number}',
            'vehicle': VehicleSerializer(vehicle).data
        })
        
    except Exception as e:
        logger.error(f"Error unassigning driver from vehicle: {str(e)}")
        return Response({
            'error': 'Erreur lors de la désassignation'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_available_drivers(request):
    """Récupérer les conducteurs disponibles"""
    try:
        user = request.user
        
        if user.is_admin():
            drivers = User.objects.filter(user_type='DRIVER', current_vehicle__isnull=True)
        elif user.is_fleet_manager():
            drivers = User.objects.filter(
                user_type='DRIVER', 
                company=user.company,
                current_vehicle__isnull=True
            )
        else:
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = UserSerializer(drivers, many=True)
        return Response({
            'drivers': serializer.data,
            'count': drivers.count()
        })
        
    except Exception as e:
        logger.error(f"Error getting available drivers: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des conducteurs disponibles'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_available_vehicles(request):
    """Récupérer les véhicules disponibles"""
    try:
        user = request.user
        
        if user.is_admin():
            vehicles = Vehicle.objects.filter(current_driver__isnull=True)
        elif user.is_fleet_manager():
            vehicles = Vehicle.objects.filter(
                company=user.company,
                current_driver__isnull=True
            )
        else:
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = VehicleSerializer(vehicles, many=True)
        return Response({
            'vehicles': serializer.data,
            'count': vehicles.count()
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"Error getting available vehicles: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des véhicules disponibles'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def update_vehicle_position(request):
    """Mettre à jour la position d'un véhicule"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent mettre à jour la position'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if not hasattr(user, 'current_vehicle') or not user.current_vehicle:
            return Response({
                'error': 'Aucun véhicule assigné'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        
        if not latitude or not longitude:
            return Response({
                'error': 'Latitude et longitude requises'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            lat = float(latitude)
            lng = float(longitude)
        except (ValueError, TypeError):
            return Response({
                'error': 'Valeurs de coordonnées invalides'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        vehicle = user.current_vehicle
        vehicle.current_latitude = lat
        vehicle.current_longitude = lng
        vehicle.last_location_update = timezone.now()
        vehicle.save()
        
        return Response({
            'success': True,
            'message': 'Position mise à jour',
            'position': {
                'latitude': vehicle.current_latitude,
                'longitude': vehicle.current_longitude,
                'timestamp': vehicle.last_location_update.isoformat()
            }
        })
        
    except Exception as e:
        logger.error(f"Error updating vehicle position: {str(e)}")
        return Response({
            'error': 'Erreur lors de la mise à jour de la position'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_vehicle_positions(request):
    """Récupérer les positions de tous les véhicules"""
    try:
        user = request.user
        
        if user.is_admin():
            vehicles = Vehicle.objects.filter(
                current_latitude__isnull=False,
                current_longitude__isnull=False
            )
        elif user.is_fleet_manager():
            vehicles = Vehicle.objects.filter(
                company=user.company,
                current_latitude__isnull=False,
                current_longitude__isnull=False
            )
        else:
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        positions = []
        for vehicle in vehicles:
            positions.append({
                'vehicle_id': vehicle.id,
                'vehicle_number': vehicle.vehicle_number,
                'latitude': vehicle.current_latitude,
                'longitude': vehicle.current_longitude,
                'last_update': vehicle.last_location_update.isoformat() if vehicle.last_location_update else None,
                'operational_status': vehicle.operational_status,
                'driver': vehicle.current_driver.get_full_name() if vehicle.current_driver else None
            })
        
        return Response({
            'positions': positions,
            'count': len(positions)
        })
        
    except Exception as e:
        logger.error(f"Error getting vehicle positions: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des positions'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_driver_current_trip(request):
    """Récupérer le voyage actuel du conducteur"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs ont des voyages'
            }, status=status.HTTP_403_FORBIDDEN)
        
        current_trip = Trip.objects.filter(
            driver=user,
            status='IN_PROGRESS'
        ).first()
        
        if current_trip:
            serializer = TripSerializer(current_trip)
            return Response({
                'has_current_trip': True,
                'trip': serializer.data
            })
        else:
            return Response({
                'has_current_trip': False,
                'trip': None
            })
        
    except Exception as e:
        logger.error(f"Error getting driver current trip: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération du voyage actuel'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_driver_trips(request):
    """Récupérer tous les voyages assignés au conducteur connecté"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs ont des voyages'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Récupérer tous les voyages du conducteur
        trips = Trip.objects.filter(driver=user).select_related(
            'vehicle', 'created_by', 'updated_by'
        ).order_by('-planned_departure')
        
        # Filtres optionnels
        status_filter = request.GET.get('status')
        if status_filter:
            trips = trips.filter(status=status_filter)
        
        serializer = TripSerializer(trips, many=True)
        return Response(serializer.data)
        
    except Exception as e:
        logger.error(f"Error getting driver trips: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des voyages'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def start_trip(request, trip_id):
    """Démarrer un voyage"""
    try:
        user = request.user
        
        try:
            trip = Trip.objects.get(id=trip_id)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier les permissions
        if not (user.is_admin() or trip.driver == user or 
                (user.is_fleet_manager() and trip.driver.company == user.company)):
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if trip.status != 'PLANNED':
            return Response({
                'error': f'Impossible de démarrer le voyage avec le statut: {trip.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        trip.status = 'IN_PROGRESS'
        trip.start_time = timezone.now()
        trip.save()
        
        serializer = TripSerializer(trip)
        return Response({
            'success': True,
            'message': 'Voyage démarré avec succès',
            'trip': serializer.data
        })
        
    except Exception as e:
        logger.error(f"Error starting trip: {str(e)}")
        return Response({
            'error': 'Erreur lors du démarrage du voyage'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def complete_trip(request, trip_id):
    """Terminer un voyage"""
    try:
        user = request.user
        
        try:
            trip = Trip.objects.get(id=trip_id)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier les permissions
        if not (user.is_admin() or trip.driver == user or 
                (user.is_fleet_manager() and trip.driver.company == user.company)):
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if trip.status != 'IN_PROGRESS':
            return Response({
                'error': f'Impossible de terminer le voyage avec le statut: {trip.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        trip.status = 'COMPLETED'
        trip.end_time = timezone.now()
        trip.save()
        
        # Mettre à jour le véhicule
        if trip.vehicle and trip.vehicle.current_driver:
            vehicle = trip.vehicle
            vehicle.operational_status = 'AVAILABLE'
            vehicle.save()
        
        serializer = TripSerializer(trip)
        return Response({
            'success': True,
            'message': 'Voyage terminé avec succès',
            'trip': serializer.data
        })
        
    except Exception as e:
        logger.error(f"Error completing trip: {str(e)}")
        return Response({
            'error': 'Erreur lors de la finalisation du voyage'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def trip_statistics(request):
    """Statistiques des voyages"""
    try:
        user = request.user
        
        if user.is_admin():
            trips = Trip.objects.all()
        elif user.is_fleet_manager():
            trips = Trip.objects.filter(driver__company=user.company)
        else:
            trips = Trip.objects.filter(driver=user)
        
        # Calculer les statistiques
        stats = {
            'total_trips': trips.count(),
            'completed_trips': trips.filter(status='COMPLETED').count(),
            'in_progress_trips': trips.filter(status='IN_PROGRESS').count(),
            'planned_trips': trips.filter(status='PLANNED').count(),
            'cancelled_trips': trips.filter(status='CANCELLED').count(),
        }
        
        # Ajouter des statistiques par période
        today = timezone.now().date()
        stats['today'] = trips.filter(created_at__date=today).count()
        stats['this_week'] = trips.filter(created_at__date__gte=today - timedelta(days=7)).count()
        stats['this_month'] = trips.filter(created_at__date__gte=today.replace(day=1)).count()
        
        return Response(stats)
        
    except Exception as e:
        logger.error(f"Error getting trip statistics: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des statistiques'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def start_trip_tracking(request, trip_id):
    """Démarrer le suivi GPS d'un voyage"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent démarrer le suivi'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trip = Trip.objects.get(id=trip_id, driver=user)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé ou non assigné'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if trip.status != 'PLANNED':
            return Response({
                'error': f'Impossible de démarrer le suivi avec le statut: {trip.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Démarrer le voyage et le suivi
        trip.status = 'IN_PROGRESS'
        trip.start_time = timezone.now()
        trip.save()
        
        return Response({
            'success': True,
            'message': 'Suivi GPS démarré',
            'trip_id': str(trip.id),
            'status': trip.status
        })
        
    except Exception as e:
        logger.error(f"Error starting trip tracking: {str(e)}")
        return Response({
            'error': 'Erreur lors du démarrage du suivi'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_trip_tracking_status(request, trip_id):
    """Récupérer le statut de suivi d'un voyage"""
    try:
        user = request.user
        
        try:
            trip = Trip.objects.get(id=trip_id)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier les permissions
        if not (user.is_admin() or trip.driver == user or 
                (user.is_fleet_manager() and trip.driver.company == user.company)):
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        tracking_status = {
            'trip_id': str(trip.id),
            'status': trip.status,
            'is_active': trip.status == 'IN_PROGRESS',
            'current_position': {
                'latitude': getattr(trip, 'current_lat', None),
                'longitude': getattr(trip, 'current_lng', None)
            } if hasattr(trip, 'current_lat') and trip.current_lat and trip.current_lng else None,
            'start_time': trip.start_time.isoformat() if trip.start_time else None,
            'driver': trip.driver.get_full_name(),
            'vehicle': trip.vehicle.vehicle_number if trip.vehicle else None
        }
        
        return Response(tracking_status)
        
    except Exception as e:
        logger.error(f"Error getting trip tracking status: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération du statut de suivi'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def vehicle_statistics(request):
    """Statistiques des véhicules"""
    try:
        user = request.user
        
        # Filtrer selon les permissions
        if user.is_admin():
            vehicles = Vehicle.objects.all()
            trips = Trip.objects.all()
        elif user.is_fleet_manager():
            vehicles = Vehicle.objects.filter(company=user.company)
            trips = Trip.objects.filter(driver__company=user.company)
        else:
            vehicles = Vehicle.objects.filter(current_driver=user)
            trips = Trip.objects.filter(driver=user)
        
        # Calculer les statistiques des véhicules
        stats = {
            'total_vehicles': vehicles.count(),
            'active_vehicles': vehicles.filter(is_active=True).count(),
            'available_vehicles': vehicles.filter(operational_status='AVAILABLE').count(),
            'in_use_vehicles': vehicles.filter(operational_status='IN_USE').count(),
            'maintenance_vehicles': vehicles.filter(operational_status='MAINTENANCE').count(),
            'out_of_service_vehicles': vehicles.filter(operational_status='OUT_OF_SERVICE').count(),
            
            # Statistiques par type de véhicule
            'vehicle_types': {
                'tractor': vehicles.filter(vehicle_type='TRACTOR').count(),
                'straight_truck': vehicles.filter(vehicle_type='STRAIGHT_TRUCK').count(),
                'van': vehicles.filter(vehicle_type='VAN').count()
            },
            
            # Statistiques des voyages
            'total_trips': trips.count(),
            'active_trips': trips.filter(status='IN_PROGRESS').count(),
            'completed_trips': trips.filter(status='COMPLETED').count(),
            'planned_trips': trips.filter(status='PLANNED').count(),
            
            # Voyages récents
            'recent_trips': TripSerializer(trips.order_by('-created_at')[:5], many=True).data,
            
            # Véhicules avec positions
            'tracked_vehicles': vehicles.filter(
                current_latitude__isnull=False,
                current_longitude__isnull=False
            ).count(),
        }
        
        return Response(stats)
        
    except Exception as e:
        logger.error(f"Error getting vehicle statistics: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des statistiques des véhicules'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_vehicles(request):
    """Export des véhicules en CSV ou Excel"""
    try:
        user = request.user
        format_type = request.GET.get('format', 'csv').lower()
        
        # Filtrer selon les permissions
        if user.is_admin():
            vehicles = Vehicle.objects.select_related('company', 'current_driver').all()
        elif user.is_fleet_manager():
            vehicles = Vehicle.objects.filter(company=user.company).select_related('company', 'current_driver')
        else:
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if format_type == 'csv':
            # Export CSV
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="vehicles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # En-têtes
            writer.writerow([
                'Numéro de véhicule', 'Type', 'Marque', 'Modèle', 'Année', 
                'Plaque d\'immatriculation', 'VIN', 'Compagnie', 'Statut opérationnel',
                'Conducteur assigné', 'Email conducteur', 'Téléphone conducteur',
                'Position latitude', 'Position longitude', 'Dernière mise à jour position',
                'Actif', 'Date de création', 'Dernière modification'
            ])
            
            # Données
            for vehicle in vehicles:
                writer.writerow([
                    vehicle.vehicle_number,
                    vehicle.get_vehicle_type_display(),
                    vehicle.make,
                    vehicle.model,
                    vehicle.year,
                    vehicle.license_plate,
                    vehicle.vin,
                    vehicle.company.name if vehicle.company else '',
                    vehicle.get_operational_status_display(),
                    vehicle.current_driver.get_full_name() if vehicle.current_driver else '',
                    vehicle.current_driver.email if vehicle.current_driver else '',
                    getattr(vehicle.current_driver, 'phone_number', '') if vehicle.current_driver else '',
                    vehicle.current_latitude or '',
                    vehicle.current_longitude or '',
                    vehicle.last_location_update.strftime('%Y-%m-%d %H:%M:%S') if vehicle.last_location_update else '',
                    'Oui' if vehicle.is_active else 'Non',
                    vehicle.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    vehicle.updated_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            return response
            
        elif format_type == 'excel':
            # Export Excel
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Véhicules"
            
            # En-têtes avec style
            headers = [
                'Numéro de véhicule', 'Type', 'Marque', 'Modèle', 'Année', 
                'Plaque d\'immatriculation', 'VIN', 'Compagnie', 'Statut opérationnel',
                'Conducteur assigné', 'Email conducteur', 'Téléphone conducteur',
                'Position latitude', 'Position longitude', 'Dernière mise à jour position',
                'Actif', 'Date de création', 'Dernière modification'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            
            # Données
            for row, vehicle in enumerate(vehicles, 2):
                worksheet.cell(row=row, column=1, value=vehicle.vehicle_number)
                worksheet.cell(row=row, column=2, value=vehicle.get_vehicle_type_display())
                worksheet.cell(row=row, column=3, value=vehicle.make)
                worksheet.cell(row=row, column=4, value=vehicle.model)
                worksheet.cell(row=row, column=5, value=vehicle.year)
                worksheet.cell(row=row, column=6, value=vehicle.license_plate)
                worksheet.cell(row=row, column=7, value=vehicle.vin)
                worksheet.cell(row=row, column=8, value=vehicle.company.name if vehicle.company else '')
                worksheet.cell(row=row, column=9, value=vehicle.get_operational_status_display())
                worksheet.cell(row=row, column=10, value=vehicle.current_driver.get_full_name() if vehicle.current_driver else '')
                worksheet.cell(row=row, column=11, value=vehicle.current_driver.email if vehicle.current_driver else '')
                worksheet.cell(row=row, column=12, value=getattr(vehicle.current_driver, 'phone_number', '') if vehicle.current_driver else '')
                worksheet.cell(row=row, column=13, value=vehicle.current_latitude or '')
                worksheet.cell(row=row, column=14, value=vehicle.current_longitude or '')
                worksheet.cell(row=row, column=15, value=vehicle.last_location_update if vehicle.last_location_update else '')
                worksheet.cell(row=row, column=16, value='Oui' if vehicle.is_active else 'Non')
                worksheet.cell(row=row, column=17, value=vehicle.created_at)
                worksheet.cell(row=row, column=18, value=vehicle.updated_at)
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Créer la réponse
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="vehicles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
        
        else:
            return Response({
                'error': 'Format non supporté. Utilisez csv ou excel.'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error exporting vehicles: {str(e)}")
        return Response({
            'error': 'Erreur lors de l\'export'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def dashboard_vehicle_stats(request):
    """Statistiques spécifiques pour le dashboard admin des véhicules"""
    try:
        user = request.user
        
        # Filtrer selon les permissions
        if user.is_admin():
            vehicles = Vehicle.objects.all()
        elif user.is_fleet_manager():
            vehicles = Vehicle.objects.filter(company=user.company)
        else:
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Calculer les statistiques détaillées
        stats = {
            'total_vehicles': vehicles.count(),
            'active_vehicles': vehicles.filter(is_active=True).count(),
            'available_vehicles': vehicles.filter(operational_status='AVAILABLE').count(),
            'in_use_vehicles': vehicles.filter(operational_status='IN_USE').count(),
            'maintenance_vehicles': vehicles.filter(operational_status='MAINTENANCE').count(),
            'out_of_service_vehicles': vehicles.filter(operational_status='OUT_OF_SERVICE').count(),
            
            # Statistiques par type
            'vehicle_types': {
                'tractor': vehicles.filter(vehicle_type='TRACTOR').count(),
                'straight_truck': vehicles.filter(vehicle_type='STRAIGHT_TRUCK').count(),
                'van': vehicles.filter(vehicle_type='VAN').count()
            },
            
            # Statistiques d'assignation
            'assigned_vehicles': vehicles.filter(current_driver__isnull=False).count(),
            'unassigned_vehicles': vehicles.filter(current_driver__isnull=True).count(),
            
            # Véhicules avec position GPS
            'tracked_vehicles': vehicles.filter(
                current_latitude__isnull=False,
                current_longitude__isnull=False
            ).count(),
            
            # Statistiques par compagnie (pour les admins)
            'by_company': {},
            
            # Récents véhicules créés
            'recent_vehicles': VehicleSerializer(
                vehicles.order_by('-created_at')[:5], 
                many=True
            ).data
        }
        
        # Ajouter les stats par compagnie pour les admins
        if user.is_admin():
            from accounts.models import Company
            companies = Company.objects.all()
            for company in companies:
                company_vehicles = vehicles.filter(company=company)
                stats['by_company'][company.name] = {
                    'total': company_vehicles.count(),
                    'active': company_vehicles.filter(is_active=True).count(),
                    'available': company_vehicles.filter(operational_status='AVAILABLE').count(),
                    'in_use': company_vehicles.filter(operational_status='IN_USE').count()
                }
        
        return Response(stats)
        
    except Exception as e:
        logger.error(f"Error getting dashboard vehicle stats: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération des statistiques'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def start_pickup(request, trip_id):
    """Démarrer le ramassage (début du voyage)"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent démarrer un ramassage'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trip = Trip.objects.get(id=trip_id, driver=user)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé ou non assigné'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if trip.status != 'PLANNED':
            return Response({
                'error': f'Impossible de démarrer le ramassage avec le statut: {trip.get_status_display()}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Récupérer la position actuelle
        current_latitude = request.data.get('latitude')
        current_longitude = request.data.get('longitude')
        
        with transaction.atomic():
            # Démarrer le voyage au point de ramassage
            trip.status = 'IN_PROGRESS'
            trip.pickup_actual_time = timezone.now()
            
            if current_latitude and current_longitude:
                trip.current_lat = float(current_latitude)
                trip.current_lng = float(current_longitude)
            
            trip.save()
            
            # Créer un log ELD pour le début du ramassage
            from eld_logs.models import ELDLog
            ELDLog.objects.create(
                driver=user,
                vehicle=trip.vehicle,
                trip=trip,
                status='DRIVING',
                location_lat=trip.current_lat if hasattr(trip, 'current_lat') else None,
                location_lng=trip.current_lng if hasattr(trip, 'current_lng') else None,
                location_description=trip.pickup_address,
                notes=f'Début du ramassage - Voyage #{trip.id}'
            )
        
        return Response({
            'success': True,
            'message': 'Ramassage démarré avec succès',
            'trip': TripSerializer(trip).data,
            'next_action': 'TRANSIT',
            'instructions': 'Vous êtes maintenant en route vers la destination'
        })
        
    except Exception as e:
        logger.error(f"Erreur lors du démarrage du ramassage: {str(e)}")
        return Response({
            'error': 'Erreur lors du démarrage du ramassage'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def complete_delivery(request, trip_id):
    """Terminer la livraison (fin du voyage)"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent terminer une livraison'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trip = Trip.objects.get(id=trip_id, driver=user)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé ou non assigné'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if trip.status != 'IN_PROGRESS':
            return Response({
                'error': f'Impossible de terminer la livraison avec le statut: {trip.get_status_display()}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Récupérer la position finale
        final_latitude = request.data.get('latitude')
        final_longitude = request.data.get('longitude')
        delivery_notes = request.data.get('notes', '')
        signature = request.data.get('signature')  # Pour la signature de livraison
        
        with transaction.atomic():
            # Terminer le voyage à la livraison
            trip.status = 'COMPLETED'
            trip.delivery_actual_time = timezone.now()
            trip.end_time = timezone.now()
            
            if final_latitude and final_longitude:
                trip.current_lat = float(final_latitude)
                trip.current_lng = float(final_longitude)
            
            if delivery_notes:
                trip.driver_notes = f"{trip.driver_notes}\n\nNotes de livraison: {delivery_notes}".strip()
            
            trip.save()
            
            # Créer un log ELD pour la fin de la livraison
            from eld_logs.models import ELDLog
            ELDLog.objects.create(
                driver=user,
                vehicle=trip.vehicle,
                trip=trip,
                status='OFF_DUTY',
                location_lat=trip.current_lat if hasattr(trip, 'current_lat') else None,
                location_lng=trip.current_lng if hasattr(trip, 'current_lng') else None,
                location_description=trip.delivery_address,
                notes=f'Livraison terminée - Voyage #{trip.id}. {delivery_notes}'
            )
            
            # Mettre à jour le statut du véhicule
            if trip.vehicle:
                trip.vehicle.operational_status = 'AVAILABLE'
                trip.vehicle.save()
        
        return Response({
            'success': True,
            'message': 'Livraison terminée avec succès',
            'trip': TripSerializer(trip).data,
            'completion_summary': {
                'pickup_time': trip.pickup_actual_time.isoformat() if trip.pickup_actual_time else None,
                'delivery_time': trip.delivery_actual_time.isoformat() if trip.delivery_actual_time else None,
                'total_duration': str(trip.delivery_actual_time - trip.pickup_actual_time) if trip.pickup_actual_time and trip.delivery_actual_time else None
            }
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la finalisation de la livraison: {str(e)}")
        return Response({
            'error': 'Erreur lors de la finalisation de la livraison'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_driver_active_trip(request):
    """Récupérer le voyage actif du conducteur avec toutes les informations de navigation"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs ont des voyages actifs'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Récupérer le voyage en cours
        active_trip = Trip.objects.filter(
            driver=user,
            status='IN_PROGRESS'
        ).select_related('vehicle', 'created_by').first()
        
        if not active_trip:
            return Response({
                'has_active_trip': False,
                'message': 'Aucun voyage actif'
            })
        
        # Calculer les informations de navigation
        navigation_data = None
        if hasattr(active_trip, 'current_lat') and active_trip.current_lat:
            nominatim_service = NominatimService()
            
            # Calculer la distance restante
            if hasattr(active_trip, 'delivery_lat') and active_trip.delivery_lat:
                remaining_distance = nominatim_service.calculate_distance(
                    active_trip.current_lat, active_trip.current_lng,
                    active_trip.delivery_lat, active_trip.delivery_lng
                )
                
                navigation_data = {
                    'current_position': {
                        'lat': active_trip.current_lat,
                        'lng': active_trip.current_lng
                    },
                    'destination': {
                        'lat': active_trip.delivery_lat,
                        'lng': active_trip.delivery_lng,
                        'address': active_trip.delivery_address
                    },
                    'remaining_distance_km': round(remaining_distance, 2),
                    'is_near_destination': remaining_distance < 0.5,  # Moins de 500m
                    'progress_percentage': active_trip.progress_percentage if hasattr(active_trip, 'progress_percentage') else 0
                }
        
        return Response({
            'has_active_trip': True,
            'trip': TripSerializer(active_trip).data,
            'navigation': navigation_data,
            'workflow': {
                'current_stage': 'TRANSIT',  # PICKUP, TRANSIT, ou DELIVERY
                'pickup_completed': active_trip.pickup_actual_time is not None,
                'delivery_pending': active_trip.delivery_actual_time is None
            }
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération du voyage actif: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération du voyage actif'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def update_trip_position(request, trip_id):
    """Mettre à jour la position GPS pendant le voyage (logs ELD continus)"""
    try:
        user = request.user
        
        if not user.is_driver():
            return Response({
                'error': 'Seuls les conducteurs peuvent mettre à jour la position'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trip = Trip.objects.get(id=trip_id, driver=user, status='IN_PROGRESS')
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage actif non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        bearing = request.data.get('bearing', 0)
        speed = request.data.get('speed', 0)
        
        if not latitude or not longitude:
            return Response({
                'error': 'Latitude et longitude requises'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            lat = float(latitude)
            lng = float(longitude)
        except (ValueError, TypeError):
            return Response({
                'error': 'Coordonnées GPS invalides'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Mettre à jour la position du voyage
        trip.current_lat = lat
        trip.current_lng = lng
        trip.save()
        
        # Mettre à jour la position du véhicule
        if trip.vehicle:
            trip.vehicle.current_latitude = lat
            trip.vehicle.current_longitude = lng
            trip.vehicle.last_location_update = timezone.now()
            trip.vehicle.save()
        
        # Calculer la distance restante et l'ETA
        nominatim_service = NominatimService()
        remaining_distance = 0
        is_near_destination = False
        
        if hasattr(trip, 'delivery_lat') and trip.delivery_lat:
            remaining_distance = nominatim_service.calculate_distance(
                lat, lng, trip.delivery_lat, trip.delivery_lng
            )
            is_near_destination = remaining_distance < 0.5  # 500 mètres
        
        # Créer un log ELD continu (tracking)
        from eld_logs.models import ELDLog
        ELDLog.objects.create(
            driver=user,
            vehicle=trip.vehicle,
            trip=trip,
            status='DRIVING',
            location_lat=lat,
            location_lng=lng,
            notes=f'Position mise à jour - Distance restante: {remaining_distance:.2f} km'
        )
        
        response_data = {
            'success': True,
            'position_updated': True,
            'current_position': {
                'lat': lat,
                'lng': lng,
                'bearing': bearing,
                'speed_kmh': speed,
                'timestamp': timezone.now().isoformat()
            },
            'navigation': {
                'remaining_distance_km': round(remaining_distance, 2),
                'is_near_destination': is_near_destination,
                'can_complete_delivery': is_near_destination
            }
        }
        
        # Notifier si proche de la destination
        if is_near_destination:
            response_data['notification'] = {
                'type': 'ARRIVAL',
                'message': 'Vous êtes proche de votre destination. Vous pouvez compléter la livraison.',
                'action': 'COMPLETE_DELIVERY'
            }
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour de position: {str(e)}")
        return Response({
            'error': 'Erreur lors de la mise à jour de la position'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_trip_timeline(request, trip_id):
    """Récupérer la timeline complète du voyage avec tous les événements ELD"""
    try:
        user = request.user
        
        try:
            trip = Trip.objects.get(id=trip_id)
        except Trip.DoesNotExist:
            return Response({
                'error': 'Voyage non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier les permissions
        if not (user.is_admin() or trip.driver == user or 
                (user.is_fleet_manager() and trip.driver.company == user.company)):
            return Response({
                'error': 'Permission refusée'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Récupérer tous les logs ELD du voyage
        from eld_logs.models import ELDLog
        from eld_logs.serializers import ELDLogSerializer
        
        eld_logs = ELDLog.objects.filter(trip=trip).order_by('timestamp')
        
        timeline = {
            'trip': TripSerializer(trip).data,
            'events': [
                {
                    'timestamp': trip.created_at.isoformat(),
                    'type': 'TRIP_CREATED',
                    'description': 'Voyage créé',
                    'location': trip.pickup_address
                }
            ]
        }
        
        # Ajouter le début du ramassage
        if trip.pickup_actual_time:
            timeline['events'].append({
                'timestamp': trip.pickup_actual_time.isoformat(),
                'type': 'PICKUP_START',
                'description': 'Début du ramassage',
                'location': trip.pickup_address
            })
        
        # Ajouter tous les logs ELD (continus)
        for log in eld_logs:
            timeline['events'].append({
                'timestamp': log.timestamp.isoformat(),
                'type': f'ELD_{log.status}',
                'description': f'Statut ELD: {log.get_status_display()}',
                'location': log.location_description or f'Lat: {log.location_lat}, Lng: {log.location_lng}',
                'notes': log.notes
            })
        
        # Ajouter la fin de la livraison
        if trip.delivery_actual_time:
            timeline['events'].append({
                'timestamp': trip.delivery_actual_time.isoformat(),
                'type': 'DELIVERY_COMPLETE',
                'description': 'Livraison terminée',
                'location': trip.delivery_address
            })
        
        # Trier par timestamp
        timeline['events'].sort(key=lambda x: x['timestamp'])
        
        return Response(timeline)
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la timeline: {str(e)}")
        return Response({
            'error': 'Erreur lors de la récupération de la timeline du voyage'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
